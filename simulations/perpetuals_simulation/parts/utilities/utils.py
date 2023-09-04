import math
import pandas as pd
import numpy as np
import copy
import os
import random
from datetime import datetime, timedelta

from openpyxl.chart import LineChart, Reference, BarChart, AreaChart, series
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows


def fetch_asset_prices(assets, timestep, event):
    asset_prices = {}
    num_hr = math.floor(timestep / 60)
    minute = timestep % 60
    parent_directory = os.path.abspath(os.path.join(os.getcwd(), os.pardir))


    for asset in assets:
        file_path = os.path.join(parent_directory, 'data', f'{event}', f'{asset}-{event}-{num_hr}.csv')
        df = pd.read_csv(file_path)

        # Calculate the absolute difference between each minute and the input, and find the index of the row with the smallest difference
        closest_idx = (df['time']- minute).abs().idxmin()

        if asset.startswith("U"):
            price = df.loc[closest_idx, 'c']
            asset_prices.update({asset: [price]})
        else:
            use_spread = df.loc[closest_idx, 'use_spread']
            if use_spread == 'True':
                min_price = df.loc[closest_idx, 'minprice']
                max_price = df.loc[closest_idx, 'maxprice']
            else:
                min_price = df.loc[closest_idx, 'price']
                max_price = df.loc[closest_idx, 'price']

            problem = df.loc[closest_idx, 'problem_indicator']

            if problem == 'no problemo':
                problem = False
            else:
                problem = True

            time_stamp = df.loc[closest_idx, 't']

            asset_prices.update({asset: [max_price, min_price, problem, time_stamp]})

    return asset_prices

def initial_liquidity(asset_prices):
    target_ratios = [0.23, 0.05, 0.25, 0.3, 0.18]
    tot_val = asset_prices['BTC'][0] / target_ratios[0]
    sol = (tot_val / asset_prices['SOL'][0]) * target_ratios[1]
    eth = (tot_val / asset_prices['ETH'][0]) * target_ratios[2]
    usdc = (tot_val / asset_prices['USDC'][0]) * target_ratios[3]
    usdt = (tot_val / asset_prices['USDT'][0]) * target_ratios[4]
    return {'BTC': 1, 'ETH': eth, 'SOL': sol, 'USDC': usdc, 'USDT': usdt}

def get_asset_volatility(assets, timestep, event):
    asset_volatility = {}
    num_hr = math.floor(timestep / 60)
    minute = timestep % 60

    # Get the data directory
    parent_directory = os.path.abspath(os.path.join(os.getcwd(), os.pardir))

    for asset in assets:
        file_path = os.path.join(parent_directory, f'data', f'{event}', f'{asset}-{event}-{num_hr}.csv')
        df = pd.read_csv(file_path) 
        try:
            closest_idx = (df['time']- minute).abs().idxmin()
            close_prices = df.loc[closest_idx-20:closest_idx, 'price']
            price_std = close_prices.std()
            price_avg = close_prices.mean()
            volatility = price_std / price_avg  # Volatility represented as average % standard deviation
        except:
            volatility = None  # If there isn't enough data to compute volatility, return None

        asset_volatility.update({asset: volatility})
    return asset_volatility

def pool_total_holdings(pool, asset_prices):
    holdings = pool['holdings']
    tvl = 0
    for asset in holdings.keys():
        tvl += holdings[asset] * asset_prices[asset][0]
    return tvl

def pool_tvl(holdings, asset_prices, minmax=0):
    tvl = 0
    for asset in holdings.keys():
        if asset.startswith('U'):
            tvl += holdings[asset] * asset_prices[asset][0]
        else:
            tvl += holdings[asset] * asset_prices[asset][minmax]
    return tvl

def get_provider_balance(provider, asset_prices):
    lp = copy.deepcopy(provider)
    total_value = sum([lp['liquidity'][asset] * asset_prices[asset][0] for asset in lp['liquidity'].keys()])
    return total_value

def calculate_interest(position_size, duration, asset, pool, rate_params):
    duration = duration/(24 * 60)
    optimal_utilization = rate_params[0]
    slope1 = rate_params[1]
    slope2 = rate_params[2]

    total_holdings = pool['holdings'][asset]
    if asset.startswith('U'):
        total_borrowed = pool['oi_short'][asset]
    else:
        total_borrowed = pool['oi_long'][asset]

    # Handle division by zero
    if total_holdings == 0:
        return 0

    current_utilization = total_borrowed / total_holdings

    if current_utilization < optimal_utilization:
        rate = (current_utilization / optimal_utilization) * slope1
    else:
        rate = slope1 + (current_utilization - optimal_utilization) / (1 - optimal_utilization) * slope2

    cumulative_interest = duration * rate
    borrow_fee_amount = cumulative_interest * position_size

    return borrow_fee_amount

def calculate_open_pnl(traders, asset_prices):
    open_pnl_long = {asset: 0 for asset in asset_prices.keys()}
    open_pnl_short = {asset: 0 for asset in asset_prices.keys()}
    for trader_id in traders.keys():
        for asset in traders[trader_id]['positions_long'].keys():
            open_pnl_long[asset] -= traders[trader_id]['positions_long'][asset]['quantity'] * (asset_prices[asset][0] - traders[trader_id]['positions_long'][asset]['entry_price'])
        for asset in traders[trader_id]['positions_short'].keys():
            open_pnl_short[asset] -= traders[trader_id]['positions_short'][asset]['quantity'] * (traders[trader_id]['positions_short'][asset]['entry_price'] - asset_prices[asset][0])

    return [open_pnl_long, open_pnl_short]

def check_for_avail(pool, token, amount):
    pool = copy.deepcopy(pool)

    if token in ['BTC', 'ETH', 'SOL']:
        avail_asset = pool['holdings'][token] - pool['oi_long'][token]
    elif token in ['USDC', 'USDT']:
        avail_asset = pool['holdings'][token] - pool['short_interest'][token]

    if avail_asset < amount:
        return -1
    
    return 0

def to_xslx(df, name):

    wb = openpyxl.Workbook()
    sheet = wb.active
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            try:
                value = float(value)
            except:
                pass
            sheet.cell(row=r_idx, column=c_idx, value=value)
    timestamps = df.shape[0]

    # Create a traction sheet
    trac_sheet = wb.create_sheet(title="Traction charts")
    # red code
    cells = [trac_sheet.cell(row=3, column=x) for x in range(21,24)]
    cells[0].value = "Key tested values"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    [setattr(cell, 'fill', red_fill) for cell in cells]
    # green code
    cells = [trac_sheet.cell(row=5, column=x) for x in range(21,24)]
    cells[0].value = "Input controlled values"
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    [setattr(cell, 'fill', green_fill) for cell in cells]
    # blue code
    cells = [trac_sheet.cell(row=7, column=x) for x in range(21,24)]
    cells[0].value = "Context values"
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    [setattr(cell, 'fill', blue_fill) for cell in cells]

    trac_sheet['A1'] = "Amount of token lps"
    values = Reference(sheet, min_col=3, min_row=3, max_col=3, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "number_of_liquidity_providers"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "LPs"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "A3")

    trac_sheet['L1'] = "Amount of traders"
    values = Reference(sheet, min_col=2, min_row=3, max_col=2, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "number_of_traders"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Trads"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "L3")

    trac_sheet['A18'] = "Number of lp tokens in the pool"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=4, min_row=3, max_col=4, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "pool_lp_tokens"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Lp tokens"
    trac_sheet.add_chart(chart, "A20")

    # Create a pool chart sheet
    pool_sheet = wb.create_sheet(title="Pool charts")

    pool_sheet['A1'] = "BTC pool size"
    values = Reference(sheet, min_col=5, min_row=3, max_col=5, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A3")

    pool_sheet['J1'] = "ETH pool size"
    values = Reference(sheet, min_col=6, min_row=3, max_col=6, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "J3")

    pool_sheet['S1'] = "SOL pool size"
    values = Reference(sheet, min_col=7, min_row=3, max_col=7, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "S3")

    pool_sheet['A18'] = "USDC pool size"
    values = Reference(sheet, min_col=8, min_row=3, max_col=8, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A20")

    pool_sheet['J18'] = "USDT pool size"
    values = Reference(sheet, min_col=9, min_row=3, max_col=9, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "J20")

    # Create a pnl sheet
    pnl_sheet = wb.create_sheet(title="PNL charts")

    pnl_sheet['A1'] = "Cumulative traders pnl"
    values = Reference(sheet, min_col=10, min_row=3, max_col=10, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "cum_pnl_traders"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "A3")

    pnl_sheet['J1'] = "Max trader pnl"
    values = Reference(sheet, min_col=11, min_row=3, max_col=11, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "max_pnl_traders"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "J3")

    pnl_sheet['S1'] = "Min trader pnl"
    values = Reference(sheet, min_col=12, min_row=3, max_col=12, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "min_pnl_traders"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "S3")

    pnl_sheet['A18'] = "BTC pool pnl"
    values = Reference(sheet, min_col=36, min_row=3, max_col=36, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "A20")

    pnl_sheet['J18'] = "ETH pool pnl"
    values = Reference(sheet, min_col=37, min_row=3, max_col=37, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "J20")

    pnl_sheet['S18'] = "SOL pool pnl"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=38, min_row=3, max_col=38, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "S20")

    # Create a oi sheet
    oi_sheet = wb.create_sheet(title="OI charts")

    oi_sheet['A1'] = "OI long BTC"
    values = Reference(sheet, min_col=13, min_row=3, max_col=13, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A3")

    oi_sheet['J1'] = "OI long ETH"
    values = Reference(sheet, min_col=14, min_row=3, max_col=14, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J3")

    oi_sheet['S1'] = "OI long SOL"
    values = Reference(sheet, min_col=15, min_row=3, max_col=15, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S3")

    oi_sheet['A18'] = "OI short BTC"
    values = Reference(sheet, min_col=16, min_row=3, max_col=16, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A20")

    oi_sheet['J18'] = "OI short ETH"
    values = Reference(sheet, min_col=17, min_row=3, max_col=17, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J20")

    oi_sheet['S18'] = "OI short SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=18, min_row=3, max_col=18, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S20")

    oi_sheet['A35'] = "Nominal exposure BTC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=39, min_row=3, max_col=39, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A37")

    oi_sheet['J35'] = "Nominal exposure ETH"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=40, min_row=3, max_col=40, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J37")

    oi_sheet['S35'] = "Nominal exposure SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=41, min_row=3, max_col=41, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S37")

    oi_sheet['A52'] = "Short interest USDC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=53, min_row=3, max_col=53, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "short_interest_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A54")

    oi_sheet['J52'] = "Short interest USDT"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=54, min_row=3, max_col=54, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "short_interest_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J54")

    oi_sheet['S52'] = "Short interest total"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=55, min_row=3, max_col=55, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "short_interest_tot"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S54")

    # Create a volume sheet
    volume_sheet = wb.create_sheet(title="Volume charts")

    volume_sheet['A1'] = "Volume BTC"
    values = Reference(sheet, min_col=19, min_row=3, max_col=19, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "A3")

    volume_sheet['J1'] = "Volume ETH"
    values = Reference(sheet, min_col=20, min_row=3, max_col=20, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "J3")

    volume_sheet['S1'] = "Volume SOL"
    values = Reference(sheet, min_col=21, min_row=3, max_col=21, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "S3")

    volume_sheet['A18'] = "Number of longs"
    values = Reference(sheet, min_col=22, min_row=3, max_col=22, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_longs"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "A20")

    volume_sheet['J18'] = "Number of shorts"
    values = Reference(sheet, min_col=23, min_row=3, max_col=23, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_shorts"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "J20")

    volume_sheet['S18'] = "Number of swaps"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=24, min_row=3, max_col=24, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_swaps"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "S20")

    volume_sheet['A35'] = "Number of liquidations"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=25, min_row=3, max_col=25, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "number_of_liquidations"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    volume_sheet.add_chart(chart, "A37")

    # Create a fees sheet
    fees_sheet = wb.create_sheet(title="Fees charts")

    fees_sheet['A1'] = "Fees collected BTC"
    values = Reference(sheet, min_col=26, min_row=3, max_col=26, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "A3")

    fees_sheet['J1'] = "Fees collected ETH"
    values = Reference(sheet, min_col=27, min_row=3, max_col=27, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "J3")

    fees_sheet['S1'] = "Fees collected SOL"
    values = Reference(sheet, min_col=28, min_row=3, max_col=28, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "S3")

    fees_sheet['A18'] = "Fees collected USDC"
    values = Reference(sheet, min_col=29, min_row=3, max_col=29, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "A20")

    fees_sheet['J18'] = "Fees collected USDT"
    values = Reference(sheet, min_col=30, min_row=3, max_col=30, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "J20")

    fees_sheet['S18'] = "Treasury balance BTC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=31, min_row=3, max_col=31, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "S20")

    fees_sheet['A35'] = "Treasury balance ETH"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=32, min_row=3, max_col=32, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A37")

    fees_sheet['J35'] = "Treasury balance SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=33, min_row=3, max_col=33, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "J37")

    fees_sheet['S35'] = "Treasury balance USDC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=34, min_row=3, max_col=34, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "S37")

    fees_sheet['A52'] = "Treasury balance USDT"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=35, min_row=3, max_col=35, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A54")

    # Create a raios sheet
    fees_sheet = wb.create_sheet(title="Ratios charts")

    fees_sheet['A1'] = "Pool TVL"
    values = Reference(sheet, min_col=42, min_row=3, max_col=42, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "TVL"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A3")

    fees_sheet['J1'] = "Pool ratio BTC"
    values = Reference(sheet, min_col=43, min_row=3, max_col=43, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_perc_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    fees_sheet.add_chart(chart, "J3")

    fees_sheet['S1'] = "Pool ratio ETH"
    values = Reference(sheet, min_col=44, min_row=3, max_col=44, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_perc_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    fees_sheet.add_chart(chart, "S3")

    fees_sheet['A18'] = "Pool ratio SOL"
    values = Reference(sheet, min_col=45, min_row=3, max_col=45, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_perc_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    fees_sheet.add_chart(chart, "A20")

    fees_sheet['J18'] = "Pool ratio USDC"
    values = Reference(sheet, min_col=46, min_row=3, max_col=46, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_perc_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    fees_sheet.add_chart(chart, "J20")

    fees_sheet['S18'] = "Pool ratio USDT"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=47, min_row=3, max_col=47, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_perc_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    fees_sheet.add_chart(chart, "S20")

    # Create a lp sheet
    fees_sheet = wb.create_sheet(title="LP charts")

    fees_sheet['A1'] = "LP balance BTC"
    values = Reference(sheet, min_col=48, min_row=3, max_col=48, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "lp_bal_btc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A3")

    fees_sheet['J1'] = "LP balance ETH"
    values = Reference(sheet, min_col=49, min_row=3, max_col=49, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "lp_bal_eth"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "J3")

    fees_sheet['S1'] = "LP balance SOL"
    values = Reference(sheet, min_col=50, min_row=3, max_col=50, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "lp_bal_sol"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "S3")

    fees_sheet['A18'] = "LP balance USDC"
    values = Reference(sheet, min_col=51, min_row=3, max_col=51, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "lp_bal_usdc"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A21")

    fees_sheet['J18'] = "LP balance USDT"
    values = Reference(sheet, min_col=52, min_row=3, max_col=52, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "lp_bal_usdt"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "J21")

    # Create a contract oi sheet
    contract_oi_sheet = wb.create_sheet(title="Contract oi charts")

    contract_oi_sheet['A1'] = "Contract OI BTC long"
    values = Reference(sheet, min_col=56, min_row=3, max_col=56, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "A3")

    contract_oi_sheet['J1'] = "Contract OI ETH long"
    values = Reference(sheet, min_col=58, min_row=3, max_col=58, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "J3")

    contract_oi_sheet['S1'] = "Contract OI SOL long"
    values = Reference(sheet, min_col=60, min_row=3, max_col=60, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "S3")

    contract_oi_sheet['A18'] = "Contract OI BTC short"
    values = Reference(sheet, min_col=57, min_row=3, max_col=57, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_sheet.add_chart(chart, "A20")

    contract_oi_sheet['J18'] = "Contract OI ETH short"
    values = Reference(sheet, min_col=59, min_row=3, max_col=59, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_sheet.add_chart(chart, "J20")

    contract_oi_sheet['S18'] = "Contract OI SOL short"
    values = Reference(sheet, min_col=61, min_row=3, max_col=61, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_sheet.add_chart(chart, "S20")

    contract_oi_sheet['A35'] = "Contract OI BTC collateral"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=68, min_row=3, max_col=68, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_collateral"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "A37")

    contract_oi_sheet['J35'] = "Contract OI ETH collateral"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=69, min_row=3, max_col=69, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_collateral"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "J37")

    contract_oi_sheet['S35'] = "Contract OI SOL collateral"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=70, min_row=3, max_col=70, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_collateral"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_sheet.add_chart(chart, "S37")

    # Create a contract oi price sheet
    contract_oi_price_sheet = wb.create_sheet(title="Contract weighted price charts")

    contract_oi_price_sheet['A1'] = "Contract OI BTC long weighted price"
    values = Reference(sheet, min_col=62, min_row=3, max_col=62, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_weighted_price_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "A3")

    contract_oi_price_sheet['J1'] = "Contract OI ETH long weighted price"
    values = Reference(sheet, min_col=64, min_row=3, max_col=64, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_weighted_price_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "J3")

    contract_oi_price_sheet['S1'] = "Contract OI SOL long weighted price"
    values = Reference(sheet, min_col=66, min_row=3, max_col=66, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_weighted_price_long"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "S3")

    contract_oi_price_sheet['A18'] = "Contract OI BTC short weighted price"
    values = Reference(sheet, min_col=63, min_row=3, max_col=63, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_weighted_price_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_price_sheet.add_chart(chart, "A20")

    contract_oi_price_sheet['J18'] = "Contract OI ETH short weighted price"
    values = Reference(sheet, min_col=65, min_row=3, max_col=65, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_weighted_price_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_price_sheet.add_chart(chart, "J20")

    contract_oi_price_sheet['S18'] = "Contract OI SOL short weighted price"
    values = Reference(sheet, min_col=67, min_row=3, max_col=67, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_weighted_price_short"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    contract_oi_price_sheet.add_chart(chart, "S20")

    contract_oi_price_sheet['A35'] = "Contract OI BTC collateral weighted price"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=71, min_row=3, max_col=71, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_btc_weighted_collateral_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "A37")

    contract_oi_price_sheet['J35'] = "Contract OI ETH collateral weighted price"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=72, min_row=3, max_col=72, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_eth_weighted_collateral_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "J37")

    contract_oi_price_sheet['S35'] = "Contract OI SOL collateral weighted price"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=73, min_row=3, max_col=73, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "contract_oi_sol_weighted_collateral_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    contract_oi_price_sheet.add_chart(chart, "S37")
    
    # Create a asset pricessheet
    asset_prices_sheet = wb.create_sheet(title="Asset price charts")

    asset_prices_sheet['A1'] = "BTC price"
    values = Reference(sheet, min_col=74, min_row=3, max_col=74, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "btc_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    asset_prices_sheet.add_chart(chart, "A3")

    asset_prices_sheet['J1'] = "ETH price"
    values = Reference(sheet, min_col=75, min_row=3, max_col=75, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "eth_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    asset_prices_sheet.add_chart(chart, "J3")

    asset_prices_sheet['S1'] = "SOL price"
    values = Reference(sheet, min_col=76, min_row=3, max_col=76, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "sol_price"
    chart.x_axis.title = "Min"
    chart.y_axis.title = "USD"
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    asset_prices_sheet.add_chart(chart, "S3")

    wb.save(f'{name}.xlsx')
