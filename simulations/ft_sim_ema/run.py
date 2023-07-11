import pandas as pd
from parts.utilities.utils import * 
from cadCAD.engine import ExecutionMode, ExecutionContext,Executor
from config import exp
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from openpyxl.chart import LineChart, Reference, BarChart, AreaChart, series
from openpyxl.styles import PatternFill


def run():
    '''
    Definition:
    Run simulation
    '''
    # Single
    exec_mode = ExecutionMode()
    local_mode_ctx = ExecutionContext(context=exec_mode.local_mode)

    simulation = Executor(exec_context=local_mode_ctx, configs=exp.configs)
    raw_system_events, tensor_field, sessions = simulation.execute()
    # Result System Events DataFrame
    df = pd.DataFrame(raw_system_events)
 
    return df

def postprocessing(df):
    json_data = df.to_json(orient='records', indent=4)

    with open('data.json', 'w') as file:
        file.write(json_data)
    
    # Pool balances and ratios
    # Amount of traders and providers
    # Cumulative PnL for traders, max and min PnL for traders (per market)
    # Cumulative apy for liquidity providers (per market)
    # Open interest (long and short supplies per market)
    # Volumes
    # Number of liquidations
    # Fees collected
    # Treasury balance

    # track pnl of the pool as metric
    # Initialize an empty list to store each timestep data
    data = []
    # Loop through each row in the dataframe
    for _, row in df.iterrows():
        traders = row['traders']
        liquidity_providers = row['liquidity_providers']
        pools = row['pools']
        # treasury = row['treasury']
        liquidations = row['liquidations']
        num_of_longs = row['num_of_longs']
        num_of_shorts = row['num_of_shorts']
        num_of_swaps = row['num_of_swaps']

        nominal_exposure_btc = 0
        nominal_exposure_eth = 0
        nominal_exposure_sol = 0
        for trader in traders.values():
            # print(trader)
            if 'BTC' in trader['positions_long'] and 'BTC' in trader['positions_short']:
                # print(f"nom exp btc {trader['positions_long']['BTC']['quantity']} + {trader['positions_short']['BTC']['quantity']}")
                nominal_exposure_btc += trader['positions_long']['BTC']['quantity'] - trader['positions_short']['BTC']['quantity']
            elif 'BTC' in trader['positions_long']: 
                nominal_exposure_btc += trader['positions_long']['BTC']['quantity']
            elif 'BTC' in trader['positions_short']:
                nominal_exposure_btc -= trader['positions_short']['BTC']['quantity']
            if 'ETH' in trader['positions_long'] and 'ETH' in trader['positions_short']:
                # print(f"nom exp eth {trader['positions_long']['ETH']['quantity']} + {trader['positions_short']['ETH']['quantity']}")
                nominal_exposure_eth += trader['positions_long']['ETH']['quantity'] - trader['positions_short']['ETH']['quantity']
            elif 'ETH' in trader['positions_long']:
                nominal_exposure_eth += trader['positions_long']['ETH']['quantity']
            elif 'ETH' in trader['positions_short']:
                nominal_exposure_eth -= trader['positions_short']['ETH']['quantity']
            if 'SOL' in trader['positions_long'] and 'SOL' in trader['positions_short']:
                # print(f"nom exp sol {trader['positions_long']['SOL']['quantity']} + {trader['positions_short']['SOL']['quantity']}")
                nominal_exposure_sol += trader['positions_long']['SOL']['quantity'] - trader['positions_short']['SOL']['quantity']
            elif 'SOL' in trader['positions_long']:
                nominal_exposure_sol += trader['positions_long']['SOL']['quantity']
            elif 'SOL' in trader['positions_short']:
                nominal_exposure_sol -= trader['positions_short']['SOL']['quantity']

        # Generate data for each row
        timestep_data = {
            'number_of_traders': len(traders),
            'number_of_liquidity_providers': len(liquidity_providers),
            'pool_lp_tokens': pools[0]['lp_shares'],
            'pool_balance_btc': pools[0]['holdings']['BTC'],
            'pool_balance_eth': pools[0]['holdings']['ETH'],
            'pool_balance_sol': pools[0]['holdings']['SOL'],
            'pool_balance_usdc': pools[0]['holdings']['USDC'],
            'pool_balance_usdt': pools[0]['holdings']['USDT'],
            'cum_pnl_traders': sum(trader['PnL'] for trader in traders.values()),
            'max_pnl_traders': max(trader['PnL'] for trader in traders.values()),
            'min_pnl_traders': min(trader['PnL'] for trader in traders.values()),
            # 'cum_apy_providers': sum(lp['yield'] for lp in liquidity_providers.values()),  # Assuming each LP has a 'yield' key
            'oi_long_btc': pools[0]['oi_long']['BTC'],
            'oi_long_eth': pools[0]['oi_long']['ETH'],
            'oi_long_sol': pools[0]['oi_long']['SOL'],
            'oi_short_btc': pools[0]['oi_short']['BTC'],
            'oi_short_eth': pools[0]['oi_short']['ETH'],
            'oi_short_sol': pools[0]['oi_short']['SOL'],
            'volume_btc': pools[0]['volume']['BTC'],
            'volume_eth': pools[0]['volume']['ETH'],
            'volume_sol': pools[0]['volume']['SOL'],
            'num_of_longs': num_of_longs,
            'num_of_shorts': num_of_shorts,
            'num_of_swaps': num_of_swaps,
            'number_of_liquidations': liquidations,
            'fees_collected_btc': pools[0]['total_fees_collected']['BTC'],
            'fees_collected_eth': pools[0]['total_fees_collected']['ETH'],
            'fees_collected_sol': pools[0]['total_fees_collected']['SOL'],
            'fees_collected_usdc': pools[0]['total_fees_collected']['USDC'],
            'fees_collected_usdt': pools[0]['total_fees_collected']['USDT'],
            'treasury_balance_btc': liquidity_providers['genesis']['liquidity']['BTC'] + liquidity_providers['genesis']['funds']['BTC'],
            'treasury_balance_eth': liquidity_providers['genesis']['liquidity']['ETH'] + liquidity_providers['genesis']['funds']['ETH'],
            'treasury_balance_sol': liquidity_providers['genesis']['liquidity']['SOL'] + liquidity_providers['genesis']['funds']['SOL'],
            'treasury_balance_usdc': liquidity_providers['genesis']['liquidity']['USDC'] + liquidity_providers['genesis']['funds']['USDC'],
            'treasury_balance_usdt': liquidity_providers['genesis']['liquidity']['USDT'] + liquidity_providers['genesis']['funds']['USDT'],
            'pool_open_pnl_btc': pools[0]['open_pnl_long']['BTC'] + pools[0]['open_pnl_short']['BTC'], 
            'pool_open_pnl_eth': pools[0]['open_pnl_long']['ETH'] + pools[0]['open_pnl_short']['ETH'],
            'pool_open_pnl_sol': pools[0]['open_pnl_long']['SOL'] + pools[0]['open_pnl_short']['SOL'],
            'nominal_exposure_btc': nominal_exposure_btc,
            'nominal_exposure_eth': nominal_exposure_eth,
            'nominal_exposure_sol': nominal_exposure_sol,
        }
        
        # Append the timestep data to the list
        data.append(timestep_data)

    # Convert the list of timestep data into a DataFrame
    data_df = pd.DataFrame(data)

    return data_df

def to_xslx(df, name):

    wb = openpyxl.Workbook()
    sheet = wb.active
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            try:
                value = float(value)
            except:
                pass
            sheet.cell(row=r_idx, column=c_idx, value=value)
    timestamps = df.shape[0]

    # Create a traction sheet
    trac_sheet = wb.create_sheet(title="Traction charts")
    # red code
    cells = [trac_sheet.cell(row=3, column=x) for x in range(21,24)]
    cells[0].value = "Key tested values"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    [setattr(cell, 'fill', red_fill) for cell in cells]
    # green code
    cells = [trac_sheet.cell(row=5, column=x) for x in range(21,24)]
    cells[0].value = "Input controlled values"
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    [setattr(cell, 'fill', green_fill) for cell in cells]
    # blue code
    cells = [trac_sheet.cell(row=7, column=x) for x in range(21,24)]
    cells[0].value = "Context values"
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    [setattr(cell, 'fill', blue_fill) for cell in cells]

    trac_sheet['A1'] = "Amount of token lps"
    values = Reference(sheet, min_col=3, min_row=3, max_col=3, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "number_of_liquidity_providers"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "LPs"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "A3")

    trac_sheet['L1'] = "Amount of traders"
    values = Reference(sheet, min_col=2, min_row=3, max_col=2, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "number_of_traders"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Trads"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    trac_sheet.add_chart(chart, "L3")

    trac_sheet['A18'] = "Number of lp tokens in the pool"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=4, min_row=3, max_col=4, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "pool_lp_tokens"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Lp tokens"
    trac_sheet.add_chart(chart, "A20")

    # Create a pool chart sheet
    pool_sheet = wb.create_sheet(title="pool charts")

    pool_sheet['A1'] = "BTC pool size"
    values = Reference(sheet, min_col=5, min_row=3, max_col=5, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A3")

    pool_sheet['J1'] = "ETH pool size"
    values = Reference(sheet, min_col=6, min_row=3, max_col=6, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "J3")

    pool_sheet['S1'] = "SOL pool size"
    values = Reference(sheet, min_col=7, min_row=3, max_col=7, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "S3")

    pool_sheet['A18'] = "USDC pool size"
    values = Reference(sheet, min_col=8, min_row=3, max_col=8, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "A20")

    pool_sheet['J18'] = "USDT pool size"
    values = Reference(sheet, min_col=9, min_row=3, max_col=9, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "pool_balance_usdt"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pool_sheet.add_chart(chart, "J20")

    # Create a pnl sheet
    pnl_sheet = wb.create_sheet(title="pnl charts")

    pnl_sheet['A1'] = "Cumulative traders pnl"
    values = Reference(sheet, min_col=10, min_row=3, max_col=10, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "cum_pnl_traders"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "A3")

    pnl_sheet['J1'] = "Max trader pnl"
    values = Reference(sheet, min_col=11, min_row=3, max_col=11, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "max_pnl_traders"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "J3")

    pnl_sheet['S1'] = "Min trader pnl"
    values = Reference(sheet, min_col=12, min_row=3, max_col=12, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "min_pnl_traders"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    pnl_sheet.add_chart(chart, "S3")

    pnl_sheet['A18'] = "BTC pool pnl"
    values = Reference(sheet, min_col=36, min_row=3, max_col=36, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "A20")

    pnl_sheet['J18'] = "ETH pool pnl"
    values = Reference(sheet, min_col=37, min_row=3, max_col=37, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "J20")

    pnl_sheet['S18'] = "SOL pool pnl"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=38, min_row=3, max_col=38, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "pool_pnl_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USD"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    pnl_sheet.add_chart(chart, "S20")

    # Create a oi sheet
    oi_sheet = wb.create_sheet(title="oi charts")

    oi_sheet['A1'] = "OI long BTC"
    values = Reference(sheet, min_col=13, min_row=3, max_col=13, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A3")

    oi_sheet['J1'] = "OI long ETH"
    values = Reference(sheet, min_col=14, min_row=3, max_col=14, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J3")

    oi_sheet['S1'] = "OI long SOL"
    values = Reference(sheet, min_col=15, min_row=3, max_col=15, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_long_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S3")

    oi_sheet['A18'] = "OI short BTC"
    values = Reference(sheet, min_col=16, min_row=3, max_col=16, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A20")

    oi_sheet['J18'] = "OI short ETH"
    values = Reference(sheet, min_col=17, min_row=3, max_col=17, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J20")

    oi_sheet['S18'] = "OI short SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=18, min_row=3, max_col=18, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "oi_short_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S20")

    oi_sheet['A35'] = "Nominal exposure BTC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=39, min_row=3, max_col=39, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "A37")

    oi_sheet['J35'] = "Nominal exposure ETH"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=40, min_row=3, max_col=40, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "J37")

    oi_sheet['S35'] = "Nominal exposure SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=41, min_row=3, max_col=41, max_row=timestamps)
    chart = LineChart()
    chart.add_data(values)
    chart.title = "nominal_exposure_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    oi_sheet.add_chart(chart, "S37")

    # Create a volume sheet
    volume_sheet = wb.create_sheet(title="volume charts")

    volume_sheet['A1'] = "Volume BTC"
    values = Reference(sheet, min_col=19, min_row=3, max_col=19, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "A3")

    volume_sheet['J1'] = "Volume ETH"
    values = Reference(sheet, min_col=20, min_row=3, max_col=20, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "J3")

    volume_sheet['S1'] = "Volume SOL"
    values = Reference(sheet, min_col=21, min_row=3, max_col=21, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "volume_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "0000FF"
    s.graphicalProperties.solidFill = "0000FF"
    volume_sheet.add_chart(chart, "S3")

    volume_sheet['A18'] = "Number of longs"
    values = Reference(sheet, min_col=22, min_row=3, max_col=22, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_longs"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "A20")

    volume_sheet['J18'] = "Number of shorts"
    values = Reference(sheet, min_col=23, min_row=3, max_col=23, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_shorts"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "J20")

    volume_sheet['S18'] = "Number of swaps"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=24, min_row=3, max_col=24, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "num_of_swaps"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    volume_sheet.add_chart(chart, "S20")

    volume_sheet['A35'] = "Number of liquidations"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=25, min_row=3, max_col=25, max_row=timestamps)
    chart = BarChart()
    chart.add_data(values)
    chart.title = "number_of_liquidations"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "Amt"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    volume_sheet.add_chart(chart, "A37")

    # Create a fees sheet
    fees_sheet = wb.create_sheet(title="Fees charts")

    fees_sheet['A1'] = "Fees collected BTC"
    values = Reference(sheet, min_col=26, min_row=3, max_col=26, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "A3")

    fees_sheet['J1'] = "Fees collected ETH"
    values = Reference(sheet, min_col=27, min_row=3, max_col=27, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "J3")

    fees_sheet['S1'] = "Fees collected SOL"
    values = Reference(sheet, min_col=28, min_row=3, max_col=28, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_sol"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "S3")

    fees_sheet['A18'] = "Fees collected USDC"
    values = Reference(sheet, min_col=29, min_row=3, max_col=29, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_usdc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "A20")

    fees_sheet['J18'] = "Fees collected USDT"
    values = Reference(sheet, min_col=30, min_row=3, max_col=30, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "fees_collected_usdt"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "FF0000"
    s.graphicalProperties.solidFill = "FF0000"
    fees_sheet.add_chart(chart, "J20")

    fees_sheet['S18'] = "Treasury balance BTC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=31, min_row=3, max_col=31, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_btc"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "BTC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "S20")

    fees_sheet['A35'] = "Treasury balance ETH"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=32, min_row=3, max_col=32, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_eth"
    chart.x_axis.title = "Hr"
    chart.y_axis.title = "ETH"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A37")

    fees_sheet['J35'] = "Treasury balance SOL"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=33, min_row=3, max_col=33, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_sol"
    chart.x_axis.title = "Day"
    chart.y_axis.title = "SOL"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "J37")

    fees_sheet['S35'] = "Treasury balance USDC"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=34, min_row=3, max_col=34, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_usdc"
    chart.x_axis.title = "Day"
    chart.y_axis.title = "USDC"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "S37")

    fees_sheet['A52'] = "Treasury balance USDT"
    timestamps = df.shape[0]
    values = Reference(sheet, min_col=35, min_row=3, max_col=35, max_row=timestamps)
    chart = AreaChart()
    chart.add_data(values)
    chart.title = "treasury_balance_usdt"
    chart.x_axis.title = "Day"
    chart.y_axis.title = "USDT"
    # Change bar filling and line color 
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00FF00"
    s.graphicalProperties.solidFill = "00FF00"
    fees_sheet.add_chart(chart, "A54")

    wb.save(f'{name}.xlsx')

def main():
    '''
    Definition:
    Run simulation and extract metrics
    '''
    df = run()
    df = postprocessing(df)
    to_xslx(df, 'run') 
    df = df[::3].reset_index(drop=True)
    to_xslx(df, 'run_merged') 
    return df

if __name__ == '__main__':
    main()